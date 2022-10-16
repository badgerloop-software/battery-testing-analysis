import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import random

import os
import re

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter



def get_compiled_data(data_specifier):
    #cell_tracker_file = ''
    #individual_data_wb = None

    if data_specifier == 'capacity':
        individual_data_wb = Workbook()
        individual_data_wb.remove(individual_data_wb.active)
        for i in range(1, 8):
            individual_data_wb.create_sheet('Box ' + str(i))
            individual_data_wb['Box ' + str(i)]['A1'].value = 'Cell'
            individual_data_wb['Box ' + str(i)]['B1'].value = 'OCV'
            individual_data_wb['Box ' + str(i)]['C1'].value = 'Capacity'
            individual_data_wb['Box ' + str(i)]['D1'].value = 'DCIR 1 (Idle-P1)'
            individual_data_wb['Box ' + str(i)]['E1'].value = 'DCIR 2 (P1-P2)'
        #print(individual_data_wb.get_index)
        cell_tracker_file = 'Battery Cell Tracker_Capacity_OCV.xlsx'

        cap_wb = pd.read_excel(io='CompiledDischargeFile.xlsx', sheet_name='8 Discharge')#'5 CCD')
        # TODO headers = np.array(list(cap_wb.columns.values))

        cap_data = cap_wb.to_numpy().transpose()
        print(cap_data)

        battery_info['filenames'] = cap_data[0]
        battery_info['data'] = cap_data[1]

        '''for i in range(len(cap_wb[headers[0]])):
            filename = cap_wb[headers[0]][i]
            #print(filename)
            battery_info['filenames'].append(filename)
            filename_parts = filename.split('_')
            #print(filename_parts)
            # If the box number in the filename is formatted as "box_#" instead of "box#", add the "#" element of
            # filename_parts (element 3) to the "box" element (making it "box#") and remove the "#" element
            if(len(filename_parts[2]) < 4):
                filename_parts[2] += filename_parts[3]
                del filename_parts[3]
                #print(filename)
                #print(filename_parts)
            battery_info['boxes'].append(int(filename_parts[2][3:]))
            battery_info['groups'].append(int(filename_parts[3][5:]))
            battery_info['rows'].append(int(filename_parts[5]))
            battery_info['battery_ids'].append('')
            battery_info['data'].append(cap_wb[headers[1]][i])'''
    else:
        individual_data_wb = load_workbook('Individual_Data.xlsx')
        cell_tracker_file = 'Battery Cell Tracker_DCIR.xlsx'
        battery_info['filenames'], battery_info['data'] = get_dcir_data()

        print(battery_info['filenames'])
        print(battery_info['data'][0])

    # Initialize cell tracker workbook
    tracker_wb = load_workbook(cell_tracker_file)

    # Iterate through the compiled DCIR/capacity data and
    for i in range(len(battery_info['filenames'])):
        filename_parts = battery_info['filenames'][i].split('_')
        # If the box number in the filename is formatted as "box_#" instead of "box#", add the "#" element of
        # filename_parts (element 3) to the "box" element (making it "box#") and remove the "#" element
        if (len(filename_parts[2]) < 4):
            filename_parts[2] += filename_parts[3]
            del filename_parts[3]

        box = filename_parts[2][3:]
        group = int(filename_parts[3][5:])
        row = int(filename_parts[5])

        t_ws = tracker_wb['Box ' + box]#tracker_wb['Box ' + str(battery_info['boxes'][i])]
        id_ws = individual_data_wb['Box ' + box]#individual_data_wb['Box ' + str(battery_info['boxes'][i])]

        bat_id = t_ws.cell(row=(row + 2), column=(2 * group)).value#t_ws.cell(row=(battery_info['rows'][i] + 2), column=(2 * battery_info['groups'][i])).value
        ocv = t_ws.cell(row=(row + 2), column=(2 * group + 1)).value# TODO Remove the variable and just use this where ocv is used below
                                                                    # #t_ws.cell(row=(battery_info['rows'][i] + 2), column=(2 * battery_info['groups'][i] + 1)).value

        #battery_info['battery_ids'][i] = bat_id # TODO Remove
        #battery_info['data'][i].append(ocv) # TODO Remove

        if data_specifier == 'capacity':
            next_row = len(id_ws['A'])
            #print(str(next_row))
            id_ws.cell(row=(next_row + 1), column=1).value = bat_id
            id_ws.cell(row=(next_row + 1), column=2).value = ocv #battery_info['data'][i]
            id_ws.cell(row=(next_row + 1), column=3).value = battery_info['data'][i]
        else:
            for j in range(1, len(id_ws['A']) + 1):
                if id_ws['A' + str(j)].value == bat_id:
                    id_ws.cell(row=j, column=4).value = battery_info['data'][0][i]
                    id_ws.cell(row=j, column=5).value = battery_info['data'][1][i]
                    break


    '''print('\n\n', len(battery_info['filenames']))
    print(battery_info['filenames'])
    print(battery_info['boxes'])
    print(battery_info['groups'])
    print(battery_info['rows'])
    print(battery_info['battery_ids'])
    print(battery_info['data'])

    print(battery_info['data'][battery_info['battery_ids'].index('PH190223-136284')])
    print(battery_info['filenames'][battery_info['battery_ids'].index('PH190223-136284')])
    print(battery_info['data'][battery_info['battery_ids'].index('PH190223-161254')])
    print(battery_info['filenames'][battery_info['battery_ids'].index('PH190223-161254')])

    for s in range(len(tracker_wb.sheetnames)):
        if not tracker_wb.sheetnames[s].startswith('Box'):
            print(tracker_wb.sheetnames[s].startswith('Box'))
            #print(tracker_wb.sheetnames[s].endswith(str(s)))
            print(tracker_wb.sheetnames[s])
            continue
        #print(tracker_wb.sheetnames[s])
        #if np.char.startswith(tracker_wb[tracker_wb.sheetnames[s]]['A1'].value, 'Box'):
        #    tracker_wb[tracker_wb.sheetnames[s]].
    #tracker_wb = pd.read_excel(io=cell_tracker_file, sheet_name='')  # '5 CCD')
    #print(tracker_wb)
    '''

    individual_data_wb.save('Individual_Data.xlsx')


def get_dcir_data():
    wb_idle = pd.read_excel(io='CompiledDCIR.xlsx', sheet_name='1 Idle') # Idle before pulse 1
    wb_ccd1 = pd.read_excel(io='CompiledDCIR.xlsx', sheet_name='2 CCD') # Pulse 1
    wb_ccd2 = pd.read_excel(io='CompiledDCIR.xlsx', sheet_name='3 CCD') # Pulse 2
    headers = np.array(list(wb_idle.columns.values))
    print(headers)
    headers = headers[~np.char.startswith(headers, 'Unnamed')]
    raw_data_idle = wb_idle.to_numpy().transpose()
    raw_data_pulse1 = wb_ccd1.to_numpy().transpose()
    raw_data_pulse2 = wb_ccd2.to_numpy().transpose()

    data_idle = []
    data_pulse1 = []
    data_pulse2 = []
    for i in range(len(raw_data_idle)):
        tmp2 = raw_data_pulse1[i]
        #tmp = tmp[~pd.isna(tmp)][1:]
        tmp2 = tmp2[~pd.isna(tmp2)][1:]
        #print(raw_data_pulse1[i][0])
        if len(tmp2) != 0 and (raw_data_pulse1[i][0] == 'V' or raw_data_pulse1[i][0] == 'mA'):
            tmp = raw_data_idle[i]
            tmp = tmp[~pd.isna(tmp)][-1]
            tmp3 = raw_data_pulse2[i]
            tmp3 = tmp3[~pd.isna(tmp3)][1]
            data_idle.append(tmp)
            data_pulse1.append(tmp2)
            data_pulse2.append(tmp3)
        #else:
            #print(headers[i])

    #print(data_pulse2)

    # TODO Change below to calculating DCIR1 and DCIR2 for each of the cells

    #time = []
    #voltage = []
    #mAmp = []
    #mAmpHr = []

    dcir1 = []
    dcir2 = []

    for i in range(int(len(data_idle) / 2)):#4)):
        #print(data_pulse1[i * 2] - data_idle[i * 2]) / (data_pulse1[i * 2 + 1] - data_idle[i * 2 + 1])

        # Multiply both quotients by -1000: 1000 to convert from kOhms to Ohms, -1 to account for negative signs on
        #                                   current being dropped when parsing the txt files
        dcir1.append(-1000 * (data_pulse1[i * 2][0] - data_idle[i * 2]) / (data_pulse1[i * 2 + 1][0] - data_idle[i * 2 + 1]))
        dcir2.append(-1000 * (data_pulse2[i * 2] - data_pulse1[i * 2][-1]) / (data_pulse2[i * 2 + 1] - data_pulse1[i * 2 + 1][-1]))
        #time.append(data_idle[i * 4])
        #voltage.append(data_idle[i * 2])#4 + 1])
        #mAmp.append(data_idle[i * 2 + 1])#4 + 2])
        #mAmpHr.append(data_idle[i * 4 + 3])

    '''for i in range(len(headers)):
        if(i > 10):
            break
        print(headers[len(headers) - (i + 1)])
        print(voltage[len(headers) - (i + 1)])'''

    print(len(dcir1))
    #print(dcir2)
    return headers, [dcir1, dcir2]


'''
static_properties = np.zeros((len(headers),3))
for i in range (len(headers)):
        static_properties[i][0] = random.uniform(1,2)
        static_properties[i][1] = voltage[i][0]
        static_properties[i][2] = mAmpHr[i][len(mAmpHr[i])-1]
'''

# Main function
if __name__ == '__main__':
    global battery_info # TODO This doesn't need to be a global. Just initialize it at the beginning of get_compiled_data()

    battery_info = {
        'filenames': [],
        'boxes': [],
        'groups': [],
        'rows': [],
        'battery_ids': [],
        'data': []
    }

    get_compiled_data('capacity')
    get_compiled_data('dcir')
    #test()
