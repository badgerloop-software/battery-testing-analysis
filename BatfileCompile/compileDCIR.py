import os
import re
import xlrd

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter

wb = Workbook()
# grab the active worksheet
ws = wb.active

sheetNames = ['1 Idle', '2 CCD', '3 CCD']
for i in range(3):#8):
    wb.create_sheet(sheetNames[i])
wb.remove(ws)  # remove the default work sheet

dir_list = os.listdir("BatFile")  # open folder containing the file

startCol = 1  # starting column
startRow = 2
for x in range(len(dir_list)):
    if dir_list[x] != '.DS_Store':
        file = open("BatFile/" + dir_list[x], encoding="ISO-8859-1")  # open file

    if file.name.endswith(".TXT"):
        lines = file.readlines()  # load each line into array
        j = -1
        numbers = []  # 3 sub array for different tests, each subarray contain min, v, mA, mAh
        for i in range(len(lines)):  # every time the line contains "mJ" start a new block
            if "mJ" in lines[i]:
                j += 1
                numbers.append([])
            elif j != -1:
                filtered = list(
                    map(float,
                        re.findall("[0-9.]+(?![0-9]\))", lines[i])[1:]))  # extract the numbers in the line using regex
                if len(filtered) != 0:
                    numbers[j].append(filtered)

        for k in range(len(numbers)):
            if len(numbers) == 3:
                wb.active = wb[sheetNames[k]]  # switch excel work book
            #else:
            #    wb.active = wb[sheetNames[k]]  # switch excel work book
            ws = wb.active
            ws[get_column_letter(startCol) + '1'] = os.path.basename(file.name).split('.')[0]
            ws[get_column_letter(startCol)+'2'] = 'min'
            ws[get_column_letter(startCol+1) + '2'] = 'V'
            ws[get_column_letter(startCol+2) + '2'] = 'mA'
            ws[get_column_letter(startCol+3) + '2'] = 'mAh'
            for i in range(len(numbers[k])):
                for j in range(len(numbers[k][i])):
                    colLet = get_column_letter(startCol + j)
                    pos = colLet + str(i + 3)  # get the position in excel sheet
                    ws[pos] = numbers[k][i][j]

        if len(numbers)<3:#7:
            print(file.name + "     # of data: " + str(len(numbers)))

        startCol += 5
        startRow += 1

wb.save("CompiledDCIR.xlsx")
