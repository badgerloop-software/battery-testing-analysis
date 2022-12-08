import sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook


def sort_and_partition_data(sort_expression, sort_first, sort_second, n_groups, group_size):
    """
    Sort battery cells using sort_expression and partition the sorted array into n_groups groups of size group_size

    :param sort_expression: A lambda expression for sorting the array of battery cells/cell information
    :param sort_first: A string describing the primary sorting parameter. Used for the name of the Excel sheet in which
                       the sorted/partitioned data will be placed
    :param sort_second: A string describing the secondary sorting parameter (or empty string if only using one sorting
                        parameter). Used for the name of the Excel sheet in which the sorted/partitioned data will be
                        placed
    :param n_groups: The number of partitions to make
    :param group_size: The size of the partitions to make
    """

    battery_info_arr = []

    wb = pd.read_excel(io='Individual_Data.xlsx', sheet_name='OCV and Capacity Filtered')

    # Put name, dcir1, dcir2, box, ocv, and capacity into an array with object elements
    names = list(wb['Cell'].to_numpy())
    dcir1s = list(wb['DCIR 1 (Idle-P1)'].to_numpy())
    dcir2s = list(wb['DCIR 2 (P1-P2)'].to_numpy())
    boxes = list(wb['Box'].to_numpy())
    ocvs = list(wb['OCV'].to_numpy())
    capacities = list(wb['Capacity'].to_numpy())

    # Get info from cell tracker spreadsheet
    box_wbs = []
    box_wbs.append(pd.read_excel(io='Battery Cell Tracker_DCIR.xlsx', sheet_name='Box 1'))
    box_wbs.append(pd.read_excel(io='Battery Cell Tracker_DCIR.xlsx', sheet_name='Box 2'))
    box_wbs.append(pd.read_excel(io='Battery Cell Tracker_DCIR.xlsx', sheet_name='Box 3'))
    box_wbs.append(pd.read_excel(io='Battery Cell Tracker_DCIR.xlsx', sheet_name='Box 4'))
    box_wbs.append(pd.read_excel(io='Battery Cell Tracker_DCIR.xlsx', sheet_name='Box 5'))
    box_wbs.append(pd.read_excel(io='Battery Cell Tracker_DCIR.xlsx', sheet_name='Box 6'))
    box_wbs.append(pd.read_excel(io='Battery Cell Tracker_DCIR.xlsx', sheet_name='Box 7'))
    
    # Tracks the cell IDs within each group (individual lists) within each box
    boxTracking = {
        'box1': [],
        'box2': [],
        'box3': [],
        'box4': [],
        'box5': [],
        'box6': [],
        'box7': []
    }
    
    # Add cell tracker information (lists of cell IDs for each group within each box) to boxTracking
    for i in range(7):
        for j in range(15):
            boxTracking['box' + str(i+1)].append(list(box_wbs[i]['Group ' + str(j+1)]))
   
    # Add cell information to battery_info_arr
    for i in range(len(names)):
        cell_number = -1
        group_number = -1
        
	# Find group and cell number in tracker file
        for j in range(15):
            try:
                cell_number = 1 + boxTracking['box' + str(boxes[i])][j].index(names[i])
                # Cell was found, record its group number and stop searching
                group_number = 1 + j
                break
            except ValueError:
                # Cell wasn't found within the current group. Continue search with the next group
                pass
        
        battery_info_arr.append({
            'name': names[i],
            'dcir1': dcir1s[i],
            'dcir2': dcir2s[i],
            'box': boxes[i],
            'ocv': ocvs[i],
            'capacity': capacities[i],
            'group': group_number,
            'cell': cell_number
        })

    # Make sure there are enough cells
    if((n_groups * group_size) > len(names)):
        print('The number of cells expecting to be partitioned (num_modules * cells_per_group) is greater than the number of available cells')
        return

    # Filter out nan from battery_info_arr because sort doesn't like them
    filter_arr = []
    for i in range(len(battery_info_arr)):
        filter_arr.append(not np.isnan(battery_info_arr[i]['dcir1']))
    battery_info_arr = list(np.array(battery_info_arr)[filter_arr])

    # Sort the values using the sort expression passed to this function
    battery_info_arr.sort(key=sort_expression)

    # Get the number of cells to keep
    arr_size = n_groups * group_size

    # Keep only the lowest (according to sort_expression) arr_size cells in the array
    cut_down_bat_info_arr = battery_info_arr[0:arr_size]

    # Partition the sorted cells into n_groups groups
    k, m = divmod(len(cut_down_bat_info_arr), n_groups)
    partitioned_bat_info = list(cut_down_bat_info_arr[i * k + min(i, m):(i + 1) * k + min(i + 1, m)] for i in range(n_groups))

    # Create/Load workbook to put partitioned cells into
    try:
        partitioned_wb = load_workbook(filename='Partitioned_Data.xlsx')
    except FileNotFoundError:
        partitioned_wb = Workbook()
        partitioned_wb.remove(partitioned_wb.active)

    # Create/Use sheet in workbook for adding data sorted/partitioned by sort_expression
    try:
        partitioned_wb.active = partitioned_wb['Sorted by ' + sort_first + ((', ' + sort_second) if (sort_second != '') else '')]
    except KeyError:
        partitioned_wb.active = partitioned_wb.create_sheet('Sorted by ' + sort_first + ((', ' + sort_second) if (sort_second != '') else ''))

    # Add the partitioned cells and their DCIR values into the workbook
    for i in range(len(partitioned_bat_info)):
        partitioned_wb.active.cell(row=1, column=i*10 + 1).value = 'Group ' + str(i+1)
        partitioned_wb.active.cell(row=2, column=i*10 + 1).value = 'Cell ID'
        partitioned_wb.active.cell(row=2, column=i*10 + 2).value = 'Box'
        partitioned_wb.active.cell(row=2, column=i*10 + 3).value = 'Group (in Box)'
        partitioned_wb.active.cell(row=2, column=i*10 + 4).value = 'Cell Number'
        partitioned_wb.active.cell(row=2, column=i*10 + 5).value = 'OCV'
        partitioned_wb.active.cell(row=2, column=i*10 + 6).value = 'Capacity'
        partitioned_wb.active.cell(row=2, column=i*10 + 7).value = 'DCIR 1 (Idle-P1)'
        partitioned_wb.active.cell(row=2, column=i*10 + 8).value = 'DCIR 2 (P1-P2)'
        partitioned_wb.active.cell(row=2, column=i*10 + 9).value = 'Avg(DCIR1, DCIR2)'
        for j in range(len(partitioned_bat_info[i])):
            partitioned_wb.active.cell(row=j+3, column=i*10 + 1).value = partitioned_bat_info[i][j]['name']
            partitioned_wb.active.cell(row=j+3, column=i*10 + 2).value = partitioned_bat_info[i][j]['box']
            partitioned_wb.active.cell(row=j+3, column=i*10 + 3).value = partitioned_bat_info[i][j]['group']
            partitioned_wb.active.cell(row=j+3, column=i*10 + 4).value = partitioned_bat_info[i][j]['cell']
            partitioned_wb.active.cell(row=j+3, column=i*10 + 5).value = partitioned_bat_info[i][j]['ocv']
            partitioned_wb.active.cell(row=j+3, column=i*10 + 6).value = partitioned_bat_info[i][j]['capacity']
            partitioned_wb.active.cell(row=j+3, column=i*10 + 7).value = partitioned_bat_info[i][j]['dcir1']
            partitioned_wb.active.cell(row=j+3, column=i*10 + 8).value = partitioned_bat_info[i][j]['dcir2']
            partitioned_wb.active.cell(row=j+3, column=i*10 + 9).value = (partitioned_bat_info[i][j]['dcir1'] + partitioned_bat_info[i][j]['dcir2']) / 2

    partitioned_wb.save('Partitioned_Data.xlsx')


# Main function
if __name__ == '__main__':
    if(len(sys.argv) != 3):
        print('Usage: `py partitionByDCIR.py <number_of_modules> <cells_per_module>`')
    else:
        try:
            # Get number of modules and cells per module from CLAs
            numModules = int(sys.argv[1])
            cellsInModule = int(sys.argv[2])

            # Partition cells using 3 different criteria
            sort_and_partition_data(lambda x: (x['dcir1'], x['dcir2']), 'DCIR 1', 'DCIR 2', numModules, cellsInModule)
            sort_and_partition_data(lambda x: (x['dcir2'], x['dcir1']), 'DCIR 2', 'DCIR 1', numModules, cellsInModule)
            sort_and_partition_data(lambda x: (x['dcir1'] + x['dcir2']) / 2, 'avg(DCIR 1, DCIR 2)', '', numModules, cellsInModule)
        except ValueError:
            print('Usage: `py partitionByDCIR.py <number of modules> <cells per module>`')

